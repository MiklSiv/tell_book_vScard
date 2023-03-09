import openpyxl


wookbook = openpyxl.load_workbook("cont.xlsx")
worksheet = wookbook.active
Book = []

for i in range(0, worksheet.max_row):
    man = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        man.append(col[i].value)
    Book.append(man)

for i in range(len(Book)):
    for k in range(len(Book[i])):
        if Book[i][k] != None and k != 0:
            Book[i][k] = Book[i][k].replace('-', '').replace(' ', '')

Book_new = []

for i in range(len(Book)):
    D = Book[i]
    S = []
    for k in D:
        if (k != None) and (k not in S):
            S.append(k)
    Book_new.append(S)

with open('contakt.csv', 'w') as csv_file:
    for row in Book_new:
        csv_file.write(", ".join(row))
        csv_file.write('\n')

